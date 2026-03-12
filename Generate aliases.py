"""
generate_aliases.py — GameSupportTracker
=========================================
Fetches all game names from the Archipelago Google Sheet and produces
an aliases.xlsx file for manual completion.

Columns:
  A  Sheet Name          – exact name as it appears in the GSheet (read-only reference)
  B  Aliases             – comma-separated list of alternative names / acronyms
                           (pre-filled with the sheet name itself; edit as needed)
  C  Done                – put any value here once you've reviewed the row

Usage:
  python generate_aliases.py [output_path]

  output_path defaults to "aliases.xlsx" in the same directory as this script.
  If aliases.xlsx already exists, existing Aliases and Done values are preserved
  so you don't lose your work when new games are added to the sheet.
"""

import sys
import os
import csv
import io

try:
    import requests
except ImportError:
    sys.exit("requests is required: pip install requests")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


# ── Config (mirrors config.py) ─────────────────────────────────────────────────

SHEET_ID = "1iuzDTOAvdoNe8Ne8i461qGNucg5OuEoF-Ikqs8aUQZw"
TABS = {
    "Playable Worlds": "58422002",
    "Core Verified":   "1675722515",
}
SKIP_NAMES = {
    "Game", "Do not sort",
    "Headers are locked to prevent auto-sorts from being done",
    "If something is missing, leave a comment!",
}


# ── Fetch ──────────────────────────────────────────────────────────────────────

def fetch_all_games() -> list[tuple[str, str]]:
    """Return [(tab_name, game_name), ...] for all games across all tabs."""
    results = []
    headers = {"User-Agent": "GameSupportTracker-AliasGenerator/1.0"}
    for tab_name, gid in TABS.items():
        url = (f"https://docs.google.com/spreadsheets/d/{SHEET_ID}"
               f"/export?format=csv&gid={gid}")
        try:
            r = requests.get(url, timeout=20, headers=headers)
            r.raise_for_status()
        except Exception as e:
            print(f"  [WARN] Could not fetch tab '{tab_name}': {e}")
            continue

        rows = list(csv.reader(io.StringIO(r.content.decode("utf-8"))))
        for row in rows:
            if not row:
                continue
            name = row[0].strip()
            if (not name
                    or name in SKIP_NAMES
                    or len(name) > 80
                    or name.lower() in ("status", "game", "do not sort")):
                continue
            results.append((tab_name, name))

    return results


# ── Load existing aliases ──────────────────────────────────────────────────────

def load_existing(path: str) -> dict[str, tuple[str, str]]:
    """
    Read an existing aliases.xlsx and return
    {sheet_name: (aliases_str, done_str)}.
    """
    if not os.path.exists(path):
        return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        existing = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            sheet_name = str(row[0]).strip()
            aliases    = str(row[1]).strip() if row[1] else sheet_name
            done       = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            existing[sheet_name] = (aliases, done)
        return existing
    except Exception as e:
        print(f"  [WARN] Could not read existing file: {e}")
        return {}


# ── Build workbook ─────────────────────────────────────────────────────────────

def build_workbook(games: list[tuple[str, str]],
                   existing: dict[str, tuple[str, str]]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aliases"

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_font     = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill     = PatternFill("solid", fgColor="1F2937")
    hdr_align    = Alignment(horizontal="center", vertical="center")

    col_a_font   = Font(name="Arial", size=9, color="9CA3AF", italic=True)
    col_a_fill   = PatternFill("solid", fgColor="111827")

    edit_font    = Font(name="Arial", size=9, color="E5E7EB")
    edit_fill_e  = PatternFill("solid", fgColor="1A2332")   # even rows
    edit_fill_o  = PatternFill("solid", fgColor="111827")   # odd rows

    done_font    = Font(name="Arial", size=9, color="6EE7B7")
    done_fill    = PatternFill("solid", fgColor="064E3B")

    pending_font = Font(name="Arial", size=9, color="6B7280")

    tab_font     = Font(name="Arial", size=8, color="6B7280", italic=True)

    border_bot   = Border(bottom=Side(style="thin", color="374151"))

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    # ── Header ────────────────────────────────────────────────────────────────
    headers = ["Sheet Name", "Aliases (comma-separated)", "Done ✓", "Tab"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border_bot

    # ── Data rows ─────────────────────────────────────────────────────────────
    seen: set[str] = set()   # deduplicate across tabs
    data_rows: list[tuple[str, str, str, str]] = []
    for tab_name, game_name in games:
        if game_name in seen:
            continue
        seen.add(game_name)
        existing_aliases, existing_done = existing.get(
            game_name, (game_name, ""))
        data_rows.append((game_name, existing_aliases, existing_done, tab_name))

    # Sort: undone first (so work is at the top), then alpha
    data_rows.sort(key=lambda r: (bool(r[2]), r[0].lower()))

    for ri, (sheet_name, aliases, done, tab_name) in enumerate(data_rows, start=2):
        is_done = bool(done)
        row_fill_e = done_fill if is_done else edit_fill_e
        row_fill_o = edit_fill_o

        fill = row_fill_e if ri % 2 == 0 else row_fill_o

        # A — Sheet Name (locked reference, greyed)
        c_a = ws.cell(row=ri, column=1, value=sheet_name)
        c_a.font      = col_a_font
        c_a.fill      = col_a_fill if not is_done else done_fill
        c_a.alignment = left
        c_a.border    = border_bot

        # B — Aliases (main editing column)
        c_b = ws.cell(row=ri, column=2, value=aliases)
        c_b.font      = done_font if is_done else edit_font
        c_b.fill      = fill
        c_b.alignment = left
        c_b.border    = border_bot

        # C — Done
        c_c = ws.cell(row=ri, column=3, value=done if done else "")
        c_c.font      = done_font if is_done else pending_font
        c_c.fill      = done_fill if is_done else fill
        c_c.alignment = center
        c_c.border    = border_bot

        # D — Tab (informational)
        c_d = ws.cell(row=ri, column=4, value=tab_name)
        c_d.font      = tab_font
        c_d.fill      = fill
        c_d.alignment = left
        c_d.border    = border_bot

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 18

    # ── Freeze header ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Summary row at bottom ─────────────────────────────────────────────────
    total  = len(data_rows)
    n_done = sum(1 for _, _, done, _ in data_rows if done)
    ws.append([])
    summary_row = ws.max_row + 1
    summary_cell = ws.cell(row=summary_row, column=1,
                           value=f"Total: {total}  |  Done: {n_done}  |  Remaining: {total - n_done}")
    summary_cell.font = Font(name="Arial", size=9, bold=True, color="9CA3AF")

    return wb


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    out_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "aliases.xlsx")

    print("Fetching game list from Google Sheets…")
    games = fetch_all_games()
    if not games:
        sys.exit("No games fetched. Check your internet connection.")
    print(f"  → {len(games)} entries fetched ({len({n for _, n in games})} unique games)")

    print(f"Loading existing aliases from: {out_path}")
    existing = load_existing(out_path)
    if existing:
        print(f"  → {len(existing)} existing rows preserved")
    else:
        print("  → No existing file, starting fresh")

    print("Building workbook…")
    wb = build_workbook(games, existing)

    wb.save(out_path)
    print(f"\nSaved → {out_path}")

    total = len({n for _, n in games})
    n_done = sum(1 for name in {n for _, n in games}
                 if existing.get(name, ("", ""))[1])
    print(f"Progress: {n_done}/{total} rows marked as done")
    print("\nHow to use:")
    print("  1. Open aliases.xlsx")
    print("  2. Column B: add comma-separated aliases/acronyms for each game")
    print("     (the sheet name is pre-filled — leave it if no alias needed)")
    print("  3. Column C: put any value (e.g. ✓) when a row is reviewed")
    print("  4. In GST Settings, point 'Aliases file' to this file")
    print("  5. Re-run this script anytime to pick up new games from the sheet")
    print("     (your existing aliases and Done marks are preserved)")


if __name__ == "__main__":
    main()