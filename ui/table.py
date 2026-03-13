import tkinter as tk
from tkinter import ttk

from config import (
    TABS, STATUS_COLORS, STATUS_ORDER, SORT_ICONS,
    BG, BG3, ACCENT, ACCENT2, TEXT, TEXT_DIM, BORDER, GREEN, RED
)
from data import match_poptracker, is_owned_on_steam, is_owned_on_playnite
from lang.l18n import t


def build_filter_bar(parent, app):
    """Build the two-row filter bar. Returns the count label."""
    fbar = tk.Frame(parent, bg=BG3, pady=6, padx=14)
    fbar.pack(fill="x")

    # Row 1 — tabs + search + count
    r1 = tk.Frame(fbar, bg=BG3)
    r1.pack(fill="x")

    app._show_left_btn = tk.Button(
        r1, text=t("btn_show_changes"), bg=BG3, fg=TEXT_DIM,
        font=("Courier New", 8), relief="flat", cursor="hand2",
        padx=6, pady=2, activebackground="#161b22", activeforeground=TEXT,
        command=app._toggle_left_panel)
    app._show_left_btn.pack(side="left", padx=(0, 8))
    app._show_left_btn.pack_forget()

    # Tab "All Games" first
    rb_all = tk.Radiobutton(r1, text=t("tab_all_games"), variable=app._tab_var,
                            value="All Games", command=app._on_tab_change,
                            bg=BG3, fg=TEXT, selectcolor=BG3,
                            activebackground=BG3, activeforeground=ACCENT2,
                            font=("Courier New", 9, "bold"),
                            indicatoron=False, relief="flat",
                            padx=10, pady=4, cursor="hand2")
    rb_all.pack(side="left", padx=(0, 4))

    for tab in TABS.keys():
        rb = tk.Radiobutton(r1, text=tab, variable=app._tab_var,
                            value=tab, command=app._on_tab_change,
                            bg=BG3, fg=TEXT, selectcolor=BG3,
                            activebackground=BG3, activeforeground=ACCENT2,
                            font=("Courier New", 9, "bold"),
                            indicatoron=False, relief="flat",
                            padx=10, pady=4, cursor="hand2")
        rb.pack(side="left", padx=(0, 4))

    tk.Label(r1, text="🔍", bg=BG3, fg=TEXT_DIM,
             font=("Courier New", 11)).pack(side="left", padx=(16, 4))
    tk.Entry(r1, textvariable=app._filter_var,
             bg=BG, fg=TEXT, insertbackground=TEXT,
             relief="flat", font=("Courier New", 10), width=22
             ).pack(side="left", ipady=4)
    app._filter_var.trace_add("write", lambda *a: app._refresh_table())

    count_lbl = tk.Label(r1, text="", bg=BG3, fg=TEXT_DIM,
                         font=("Courier New", 9))
    count_lbl.pack(side="right", padx=8)

    # Export Excel button — same visual style as Edit Owned button
    tk.Button(
        r1, text=t("btn_export"), bg=BG3, fg=TEXT,
        font=("Courier New", 9, "bold"), relief="raised", bd=1, cursor="hand2",
        padx=10, pady=4, activebackground=BG3, activeforeground=ACCENT2,
        command=lambda: export_to_excel(app)
    ).pack(side="right", padx=(0, 4))

    # Edit owned button — same visual style as tab radiobuttons (indicatoron=False)
    app._edit_owned_btn = tk.Button(
        r1, text=t("btn_edit_owned"), bg=BG3, fg=TEXT,
        font=("Courier New", 9, "bold"), relief="raised", bd=1, cursor="hand2",
        padx=10, pady=4, activebackground=BG3, activeforeground=ACCENT2,
        command=app._toggle_edit_owned)
    app._edit_owned_btn.pack(side="right", padx=(0, 4))

    # Row 2 — dropdowns
    r2 = tk.Frame(fbar, bg=BG3)
    r2.pack(fill="x", pady=(4, 0))

    tk.Label(r2, text=t("filter_status"), bg=BG3, fg=TEXT_DIM,
             font=("Courier New", 9)).pack(side="left", padx=(0, 4))
    _status_values_all = [t("filter_all")] + list(STATUS_COLORS.keys()) + ["Core Verified"]
    app._status_combo = ttk.Combobox(r2, textvariable=app._status_filter,
                 values=_status_values_all,
                 state="readonly", width=16,
                 font=("Courier New", 9))
    app._status_combo.pack(side="left")
    app._status_filter.trace_add("write", lambda *a: app._refresh_table())

    tk.Label(r2, text=t("filter_pt"), bg=BG3, fg=TEXT_DIM,
             font=("Courier New", 9)).pack(side="left", padx=(12, 4))
    ttk.Combobox(r2, textvariable=app._pt_filter,
                 values=[t("filter_all"), t("filter_pt_yes"), t("filter_pt_no")],
                 state="readonly", width=16,
                 font=("Courier New", 9)).pack(side="left")
    app._pt_filter.trace_add("write", lambda *a: app._refresh_table())

    tk.Label(r2, text=t("filter_owned"), bg=BG3, fg=TEXT_DIM,
             font=("Courier New", 9)).pack(side="left", padx=(12, 4))
    ttk.Combobox(r2, textvariable=app._owned_filter,
                 values=[t("filter_all"), t("filter_owned_yes"), t("filter_owned_no")],
                 state="readonly", width=8,
                 font=("Courier New", 9)).pack(side="left")
    app._owned_filter.trace_add("write", lambda *a: app._refresh_table())

    return count_lbl


def build_tree(parent, app):
    """Build the Treeview with scrollbar. Returns the tree widget."""
    table_frame = tk.Frame(parent, bg=BORDER)
    table_frame.pack(fill="both", expand=True)

    tk.Frame(table_frame, bg=BORDER, height=1).pack(fill="x", side="top")

    inner = tk.Frame(table_frame, bg=BG)
    inner.pack(fill="both", expand=True)

    style = ttk.Style()
    style.theme_use("clam")
    style.configure("Custom.Treeview",
                    background=BG, foreground=TEXT,
                    rowheight=31, fieldbackground=BG,
                    borderwidth=0, font=("Courier New", 9), relief="flat")
    style.configure("Custom.Treeview.Heading",
                    background=BG3, foreground=ACCENT2,
                    font=("Courier New", 9, "bold"), relief="flat",
                    borderwidth=0, padding=(6, 6))
    style.map("Custom.Treeview",
              background=[("selected", ACCENT)],
              foreground=[("selected", "white")])
    style.map("Custom.Treeview.Heading",
              background=[("active", ACCENT)],
              foreground=[("active", "white")])

    cols = ("game", "status", "poptracker", "notes", "owned")
    tree = ttk.Treeview(inner, columns=cols, show="headings",
                        style="Custom.Treeview")

    tree.heading("game",       text=t("col_game") + SORT_ICONS[None], anchor="w",
                 command=lambda: app._on_sort_click("game"))
    tree.heading("status",     text=t("col_status") + SORT_ICONS[None], anchor="w",
                 command=lambda: app._on_sort_click("status"))
    tree.heading("poptracker", text=t("col_poptracker") + SORT_ICONS[None], anchor="w",
                 command=lambda: app._on_sort_click("poptracker"))
    tree.heading("notes",      text=t("col_notes"), anchor="w")
    tree.heading("owned",      text=t("col_owned") + SORT_ICONS[None], anchor="w",
                 command=lambda: app._on_sort_click("owned"))

    tree.column("game",       width=220, minwidth=130, stretch=False)
    tree.column("status",     width=120, minwidth=90,  stretch=False)
    tree.column("poptracker", width=100, minwidth=80,  stretch=False)
    tree.column("notes",      width=430, minwidth=160, stretch=True)
    tree.column("owned",      width=70,  minwidth=60,  stretch=False)

    vsb = ttk.Scrollbar(inner, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    tree.pack(side="left", fill="both", expand=True)
    vsb.pack(side="right", fill="y")

    for status, color in STATUS_COLORS.items():
        tree.tag_configure(status, foreground=color)
    tree.tag_configure("Other",    foreground=TEXT_DIM)
    tree.tag_configure("new",      background="#1a2e1a")
    tree.tag_configure("core_yes", foreground=GREEN)
    tree.tag_configure("core_no",  foreground=RED)
    tree.tag_configure("odd_row",  background="#0d1117")
    tree.tag_configure("even_row", background="#161b22")

    # Click handler for edit-owned mode
    tree.bind("<Button-1>",   lambda e: _on_tree_click(e, tree, app))
    # Text selection / copy support
    tree.bind("<Button-3>",   lambda e: _on_tree_right_click(e, tree, app))
    tree.bind("<Control-c>",  lambda e: _copy_tree_selection(tree, app))
    tree.bind("<Control-C>",  lambda e: _copy_tree_selection(tree, app))

    return tree


def _copy_tree_selection(tree, app):
    """Copy selected row(s) game name(s) to clipboard (Ctrl+C)."""
    selected = tree.selection()
    if not selected:
        return
    lines = []
    for iid in selected:
        vals = tree.item(iid, "values")
        if vals:
            lines.append(vals[0])  # game name
    if lines:
        app.clipboard_clear()
        app.clipboard_append("\n".join(lines))


def _on_tree_right_click(event, tree, app):
    """Show a small context menu to copy cell or row on right-click."""
    row_id = tree.identify_row(event.y)
    col    = tree.identify_column(event.x)
    if not row_id:
        return

    tree.selection_set(row_id)
    vals = tree.item(row_id, "values")
    if not vals:
        return

    col_idx = int(col.replace("#", "")) - 1
    cell_val = vals[col_idx] if 0 <= col_idx < len(vals) else ""
    row_text = "\t".join(str(v) for v in vals)

    menu = tk.Menu(app, tearoff=0, bg="#1f2937", fg="#e5e7eb",
                   activebackground="#374151", activeforeground="white",
                   font=("Courier New", 9), bd=0, relief="flat")

    if cell_val:
        menu.add_command(
            label=f'Copier "{cell_val[:40]}{"…" if len(cell_val) > 40 else ""}"',
            command=lambda: (app.clipboard_clear(), app.clipboard_append(cell_val)))
    menu.add_command(
        label="Copier la ligne",
        command=lambda: (app.clipboard_clear(), app.clipboard_append(row_text)))

    try:
        menu.tk_popup(event.x_root, event.y_root)
    finally:
        menu.grab_release()



    """Handle clicks in edit-owned mode to toggle ownership."""
    if not app._edit_owned_mode:
        return
    region = tree.identify_region(event.x, event.y)
    if region != "cell":
        return
    col = tree.identify_column(event.x)
    if col != "#5":  # owned column
        return
    row_id = tree.identify_row(event.y)
    if not row_id:
        return
    values = tree.item(row_id, "values")
    if not values:
        return
    game_name = values[0]
    app.toggle_manual_owned(game_name)
    # Update the cell display immediately
    is_owned = game_name in app._manual_owned
    new_values = list(values)
    new_values[4] = "☑" if is_owned else "☐"
    tree.item(row_id, values=new_values)


def apply_columns(tree, tab, sort_col, sort_asc):
    """Adjust column visibility for the active tab."""
    is_core = tab == "Core Verified"
    if is_core:
        tree.column("status", width=0, minwidth=0, stretch=False)
        tree.heading("status", text="")
        tree.column("game",  width=260, minwidth=140)
        tree.column("notes", width=480, minwidth=180)
    else:
        tree.column("status", width=120, minwidth=90, stretch=False)
        tree.heading("status", text=t("col_status") + SORT_ICONS[
            sort_asc if sort_col == "status" else None])
        tree.column("game",  width=220, minwidth=130)
        tree.column("notes", width=430, minwidth=160)


def update_heading_icons(tree, tab, sort_col, sort_asc):
    is_core = tab == "Core Verified"
    cols_labels = {"game": t("col_game"), "poptracker": t("col_poptracker"), "owned": t("col_owned")}
    if not is_core:
        cols_labels["status"] = t("col_status")
    for col, label in cols_labels.items():
        icon = SORT_ICONS[sort_asc] if sort_col == col else SORT_ICONS[None]
        tree.heading(col, text=label + icon)


def sort_key(item, sort_col):
    name, data, has_pt, is_owned = item[0], item[1], item[2], item[3]
    if sort_col == "game":
        return name.lower()
    elif sort_col == "status":
        return (STATUS_ORDER.get(data.get("status", ""), 99), name.lower())
    elif sort_col == "poptracker":
        return (0 if has_pt else 1, name.lower())
    elif sort_col == "owned":
        return (0 if is_owned else 1, name.lower())
    return name.lower()


def refresh_table(tree, app):
    """Repopulate the treeview from app state."""
    for item in tree.get_children():
        tree.delete(item)

    tab = app._tab_var.get()
    edit_mode = app._edit_owned_mode

    # All Games: merge all tabs
    if tab == "All Games":
        games = {}
        source_map = {}
        for tab_name in TABS.keys():
            tab_data = app._all_games.get(tab_name, {})
            if isinstance(tab_data, dict):
                for name, data in tab_data.items():
                    if name not in games:
                        games[name] = data
                        source_map[name] = tab_name
    else:
        games = app._all_games.get(tab, {})
        source_map = {name: tab for name in games}

    if not isinstance(games, dict):
        return

    query       = app._filter_var.get().lower()
    sf          = app._status_filter.get()
    pt_filt     = app._pt_filter.get()
    owned_filt  = app._owned_filter.get()
    is_all      = tab == "All Games"
    new_names   = {e[2] for e in app._changes if e[0] == "➕" and (is_all or e[1] == tab)}
    is_core     = tab == "Core Verified"

    yes_txt = t("cell_yes")
    no_txt  = t("cell_no")

    filtered = []
    for name, data in games.items():
        if not isinstance(data, dict):
            continue
        status   = data.get("status", "")
        notes    = data.get("notes",  "")
        has_pt   = match_poptracker(name, app._poptracker_set)
        # Owned = Steam OR Playnite OR manual
        is_owned = (is_owned_on_steam(name, app._steam_owned,
                                       getattr(app, "_steam_bases", None))
                    or is_owned_on_playnite(name, app._playnite_owned,
                                            getattr(app, "_playnite_bases", None))
                    or name in app._manual_owned)
        src      = source_map.get(name, "")

        if query and query not in name.lower() \
                 and query not in status.lower() \
                 and query not in notes.lower():
            continue
        if sf != "All":
            if sf == "Core Verified":
                if src != "Core Verified":
                    continue
            elif status != sf:
                continue
        if pt_filt == t("filter_pt_yes") and not has_pt:   continue
        if pt_filt == t("filter_pt_no")  and has_pt:       continue
        if owned_filt == t("filter_owned_yes") and not is_owned: continue
        if owned_filt == t("filter_owned_no")  and is_owned:     continue

        filtered.append((name, data, has_pt, is_owned, src))

    if app._sort_col is not None:
        filtered.sort(key=lambda x: sort_key(x, app._sort_col),
                      reverse=(app._sort_asc is False))
    else:
        filtered.sort(key=lambda x: x[0].lower())

    for idx, (name, data, has_pt, is_owned, src) in enumerate(filtered):
        status    = data.get("status", "")
        notes     = data.get("notes",  "")
        pt_txt    = yes_txt if has_pt   else no_txt

        if edit_mode:
            owned_txt = "☑" if is_owned else "☐"
        else:
            owned_txt = yes_txt if is_owned else no_txt

        row_is_core = (tab == "Core Verified") or (is_all and src == "Core Verified" and not status)
        if row_is_core:
            row_tag = "core_yes" if has_pt else "core_no"
            display_status = "Core Verified" if is_all else status
        else:
            row_tag = status if status in STATUS_COLORS else "Other"
            display_status = status

        stripe = "even_row" if idx % 2 == 0 else "odd_row"
        tags   = [stripe, row_tag] + (["new"] if name in new_names else [])
        tree.insert("", "end",
                    values=(name, display_status, pt_txt, notes, owned_txt),
                    tags=tags)

    app._count_lbl.config(text=t("count_label", n=len(filtered)))

    # Update status filter values based on active tab
    if hasattr(app, "_status_combo"):
        base   = [t("filter_all")] + list(STATUS_COLORS.keys())
        values = base + ["Core Verified"] if tab == "All Games" else base
        app._status_combo.config(values=values)
        if app._status_filter.get() == "Core Verified" and tab != "All Games":
            app._status_filter.set(t("filter_all"))

# ── Excel export ───────────────────────────────────────────────────────────────

def export_to_excel(app):
    """
    Export the currently visible (filtered + sorted) table rows to an .xlsx file.
    Opens a Save-As dialog; the exported sheet mirrors the active tab/filters/sort.
    Uses openpyxl if available, falls back to csv otherwise.
    """
    from tkinter import filedialog, messagebox
    import os

    tree = app._tree
    tab  = app._tab_var.get()

    # Collect current visible rows from the tree (already filtered + sorted)
    rows = []
    for iid in tree.get_children():
        vals = tree.item(iid, "values")
        if vals:
            rows.append(list(vals))

    if not rows:
        messagebox.showinfo(t("export_empty_title"), t("export_empty_msg"))
        return

    # Column headers (respect Core Verified tab which hides Status)
    is_core = tab == "Core Verified"
    if is_core:
        headers = [t("col_game"), t("col_poptracker"), t("col_notes"), t("col_owned")]
        col_indices = [0, 2, 3, 4]
    else:
        headers = [t("col_game"), t("col_status"), t("col_poptracker"), t("col_notes"), t("col_owned")]
        col_indices = [0, 1, 2, 3, 4]

    safe_tab = tab.replace("/", "-").replace("\\", "-")
    default_name = f"GST_{safe_tab}.xlsx"

    path = filedialog.asksaveasfilename(
        title=t("export_dialog_title"),
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv"), (t("export_all_files"), "*.*")],
        initialfile=default_name,
    )
    if not path:
        return

    try:
        if path.endswith(".csv"):
            _export_csv(path, headers, rows, col_indices)
        else:
            _export_xlsx(path, headers, rows, col_indices, tab)
        messagebox.showinfo(t("export_success_title"),
                            t("export_success_msg", path=os.path.basename(path)))
    except Exception as exc:
        messagebox.showerror(t("export_error_title"), str(exc))


def _export_csv(path, headers, rows, col_indices):
    import csv
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for row in rows:
            w.writerow([row[i] for i in col_indices])


def _export_xlsx(path, headers, rows, col_indices, sheet_name):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # openpyxl not installed → fallback to CSV with .xlsx extension renamed
        _export_csv(path.replace(".xlsx", ".csv"), headers, rows, col_indices)
        raise RuntimeError(
            "openpyxl n'est pas installé. Le fichier a été sauvegardé en CSV. "
            "Installez openpyxl avec: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    safe = sheet_name[:31]  # Excel sheet name max 31 chars
    ws.title = safe

    # ── Styles ──────────────────────────────────────────────────────────────
    hdr_font  = Font(name="Courier New", bold=True, color="E5E7EB", size=9)
    hdr_fill  = PatternFill("solid", fgColor="1F2937")
    cell_font = Font(name="Courier New", size=9, color="E5E7EB")
    alt_fill  = PatternFill("solid", fgColor="161B22")
    base_fill = PatternFill("solid", fgColor="0D1117")
    thin = Side(style="thin", color="374151")
    border = Border(bottom=Side(style="thin", color="374151"))

    # ── Header row ──────────────────────────────────────────────────────────
    for ci, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Data rows ───────────────────────────────────────────────────────────
    for ri, row in enumerate(rows, start=2):
        fill = base_fill if ri % 2 == 0 else alt_fill
        for ci, col_i in enumerate(col_indices, start=1):
            val  = row[col_i] if col_i < len(row) else ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = cell_font
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # ── Column widths (auto-fit to content, capped) ─────────────────────────
    for ci, header in enumerate(headers, start=1):
        col_letter = get_column_letter(ci)
        max_len = len(header)
        for row in rows:
            col_i = col_indices[ci - 1]
            val = str(row[col_i]) if col_i < len(row) else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(path)